import pandas as pd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from mpmath import *
import sympy
import cmath
import pandas
import math


# input from txt
def txt(file, wl_txt, n_txt, k_txt):
    with open(file, 'r') as file:
        space = True
        for line in file:
            if not line.isspace():
                if space:
                    buf1, buf2 = line.split('\t')
                    if buf1 != 'wl' and buf2 != 'n':
                        wl_txt.append(float(buf1))
                        n_txt.append(float(buf2))
                else:
                    buf1, buf2 = line.split('\t')
                    if buf1 != 'wl' and buf2 != 'k':
                        k_txt.append(float(buf2))
            if line.isspace():
                space = False


# input from excel
def excel(file, wl_xlsx, n_xlsx, k_xlsx):
    test = openpyxl.reader.excel.load_workbook(filename=file, data_only=True)
    test.active = 0
    sheet = test.active
    for i in range(2, sheet.max_row + 1):
        if sheet['A' + str(i)].value >= 0.1195:
            wl_xlsx.append(float(sheet['A' + str(i)].value))
            n_xlsx.append(float(sheet['B' + str(i)].value))
            k_xlsx.append(float(sheet['C' + str(i)].value))


# psi function
def psi(n, z):
    mp.dps = 6
    mp.pretty = True
    return (sqrt(pi * z / 2) * besselj(n + (1 / 2), z)).real


def psi_diff(n, z):
    mp.dps = 6
    mp.pretty = True
    return (sqrt(pi / 2) * (
            1 / (2 * sqrt(z)) * besselj(n + (1 / 2), z) + sqrt(z) * diff(lambda z: besselj(n + (1 / 2), z),
                                                                         z))).real


# eps function
def eps(n, z):
    mp.dps = 6
    mp.pretty = True
    return sqrt(pi * z / 2) * (besselj(n + (1 / 2), z) + bessely(n + (1 / 2), z) * j)


def eps_diff(n, z):
    mp.dps = 6
    mp.pretty = True
    return sqrt(pi / 2) * (1 / (2 * sqrt(z))) * (besselj(n + (1 / 2), z) + bessely(n + (1 / 2), z) * j) + sqrt(
        pi * z / 2) * (diff(lambda z: besselj(n + (1 / 2), z), z) + (diff(lambda z: bessely(n + (1 / 2), z), z)) * j)


def a(n, m, x):
    up = m * psi(n, m * x) * psi_diff(n, x) - psi(n, x) * psi_diff(n, m * x)
    down = m * psi(n, m * x) * eps_diff(n, x) - eps(n, x) * psi_diff(n, m * x)
    return up / down


def b(n, m, x):
    up = psi(n, m * x) * psi_diff(n, x) - m * psi(n, x) * psi_diff(n, m * x)
    down = psi(n, m * x) * eps_diff(n, x) - m * eps(n, x) * psi_diff(n, m * x)
    return up / down


# relative refractive index
def m(n_p, k_p, n_f, k_f):
    return (n_p + k_p * j) / (n_f + k_f * j)


# wave number
def k(n_f, l):
    return 2 * pi * n_f / l


# the parameter of diffraction
def x(n_f, l, d):
    return d * pi * n_f / l


def n_max(n_f, l, d):
    return int(2 + x(n_f, l, d) + 4 * (x(n_f, l, d) ** (1 / 3)))


# Extinction cross section
def extinction(l, n_p, k_p, n_f, k_f, d):
    s = 0
    for i in range(1, n_max(n_f, l, d) + 1):
        s += (2 * i + 1) * (a(i, m(n_p, k_p, n_f, k_f), x(n_f, l, d)) + b(i, m(n_p, k_p, n_f, k_f), x(n_f, l, d))).real
    return 2 * pi * s / (k(n_f, l) * k(n_f, l))


# Solar spectra
wl = []  # Wavelength, nm
I_0 = []  # Intensity
test = openpyxl.reader.excel.load_workbook(filename='Solar_spectra.xlsx', data_only=True)
test.active = 2
sheet = test.active
for i in range(3, 1700):
    wl.append(float(sheet['F' + str(i)].value))
    I_0.append(float(sheet['G' + str(i)].value))

# Water.xlsx
wl_water = []
n_water = []
k_water = []
excel('Water.xlsx', wl_water, n_water, k_water)

# Graphite.xlsx
wl_graphite = []
n_graphite = []
k_graphite = []
excel('Graphite.xlsx', wl_graphite, n_graphite, k_graphite)

print('Input volume fraction')
f_v = float(input())
print('Input the particle diameter, m')
d = float(input())
print('Input nodal points of the depth of the nanofluid layer, m')
depth = [float(i) for i in input().split()]

C_ext = []  # Extinction cross section
sigma_p = []  # Extinction of particles
sigma_f = []  # Extinction of liquid
ext = []  # Nanofluid extinction
for i in range(len(wl)):
    l = wl[i] * 10e-9  # Wavelength
    n_p = np.interp(wl[i] * 10e-3, wl_graphite, n_graphite)
    k_p = np.interp(wl[i] * 10e-3, wl_graphite, k_graphite)
    n_f = np.interp(wl[i] * 10e-3, wl_water, n_water)
    k_f = np.interp(wl[i] * 10e-3, wl_water, k_water)
    C = extinction(l, n_p, k_p, n_f, k_f, d)
    C_ext.append(C)
    sigma_p.append(6 * f_v * C / (pi * (d ** 3)))
    sigma_f.append(4 * pi * k_f / l)
    ext.append(6 * f_v * C / (pi * (d ** 3)) + (1 - f_v) * 4 * pi * k_f / l)

q = []
for i in range(len(depth)):
    sum = 0
    for j in range(len(wl) - 1):
        a = (I_0[j] * exp(-ext[j] * depth[i]) + I_0[j + 1] * exp(-ext[j + 1] * depth[i])) / 2  # The trapezoid procedure
        sum += a * (wl[j + 1] - wl[j])
    q.append(sum)

# Output results
my_wb = openpyxl.Workbook()  # new Workbook
my_sheet = my_wb.active
name1 = my_sheet['A1']
name1.value = "Wavelength, nm"
name2 = my_sheet['B1']
name2.value = "Extinction cross section, m^2"
name3 = my_sheet['C1']
name3.value = "Extinction of particles, m^-1"
name4 = my_sheet['D1']
name4.value = "Extinction of liquid, m^-1"
name5 = my_sheet['E1']
name5.value = "Nanofluid extinction, m^-1"
name6 = my_sheet['G1']
name6.value = "Depth, m"
name7 = my_sheet['H1']
name7.value = "Thermal flow density, W*m^-2"
for i in range(len(wl)):
    buf1 = my_sheet['A' + str(i + 2)]
    buf1.value = float(wl[i])
    buf2 = my_sheet['B' + str(i + 2)]
    buf2.value = float(C_ext[i])
    buf3 = my_sheet['C' + str(i + 2)]
    buf3.value = float(sigma_p[i])
    buf4 = my_sheet['D' + str(i + 2)]
    buf4.value = float(sigma_f[i])
    buf5 = my_sheet['E' + str(i + 2)]
    buf5.value = float(ext[i])
for i in range(len(depth)):
    buf6 = my_sheet['G' + str(i + 2)]
    buf6.value = float(depth[i])
    buf7 = my_sheet['H' + str(i + 2)]
    buf7.value = float(q[i])
my_wb.save("Result.xlsx")

# Output of spectra and graphs
spectrum, ax = plt.subplots()
ax.plot(wl, ext)
ax.grid()
plt.xlim([np.min(wl), 2000])
ax.set_xlabel('Wavelength, nm')
ax.set_ylabel('Nanofluid extinction, m^-1')
plt.show()

graph, ax = plt.subplots()
ax.plot(depth, q)
ax.grid()
ax.set_xlabel('Depth, m')
ax.set_ylabel('Thermal flow density, W*m^-2')
plt.show()